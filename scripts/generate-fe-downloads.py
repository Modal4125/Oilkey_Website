#!/usr/bin/env python3
"""Generate downloadable XLSX and PDF editions of the F@ctory Express catalogue.

Reads src/data/factory-express.ts (ja) and factory-express-en.ts (en) and
writes four files into public/downloads/. Rerun after any catalogue data
change so the downloads stay in sync with the web pages:

    python3 scripts/generate-fe-downloads.py

Requires: node on PATH, openpyxl, reportlab.
"""

import json
import re
import subprocess
import tempfile
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "public" / "downloads"

META = {
    "ja": {
        "src": ROOT / "src/data/factory-express.ts",
        "export": "factoryExpress",
        "title": "F@ctory Express 製品カタログ",
        "company": "オイルキー株式会社",
        "contact": "TEL 072-284-1711 ／ FAX 072-298-7790 ／ oilkey@oilkey.co.jp",
        "disclaimer": "Eプライスはご注文数量が1個の場合の価格（円）です。10個単位のご注文には割引価格が適用されます。価格・納期は予告なく変更される場合があります。",
        "linknote": "製品名のリンクは、OKS社（oks-germany.com）の製品ページを開きます。",
        "generated": "発行日",
        "products": "製品",
    },
    "en": {
        "src": ROOT / "src/data/factory-express-en.ts",
        "export": "factoryExpressEn",
        "title": "F@ctory Express Product Catalogue",
        "company": "Oilkey Corporation",
        "contact": "TEL +81-72-284-1711 / FAX +81-72-298-7790 / oilkey@oilkey.co.jp",
        "disclaimer": "The E-price is the unit price (JPY) for an order of one; discounted pricing applies to orders in units of ten. Prices and lead times are subject to change without notice.",
        "linknote": "Product names link to the manufacturer's product pages on oks-germany.com.",
        "generated": "Issued",
        "products": "products",
    },
}

NAVY = "02376B"
LINK_BLUE = "0B5FA5"


def load_oks_links():
    """OKS number → product URL, parsed from src/data/oks-links.ts."""
    ts = (ROOT / "src/data/oks-links.ts").read_text()
    base = re.search(r"const BASE = '([^']+)'", ts).group(1)
    return {num: base + path for num, path in re.findall(r"'(\d+)': `\$\{BASE\}([^`]+)`", ts)}


def oks_link(links, name_cell):
    m = re.match(r"OKS\s*(\d+)", name_cell)
    return (m.group(0), links.get(m.group(1))) if m else (None, None)


def load_categories(lang):
    ts = META[lang]["src"].read_text()
    ts = re.sub(r"^import type.*$", "", ts, flags=re.M)
    ts = re.sub(r"export interface FeCategory \{.*?\n\}", "", ts, flags=re.S)
    ts = ts.replace(": FeCategory[]", "")
    with tempfile.NamedTemporaryFile("w", suffix=".mjs", delete=False) as f:
        f.write(ts + f"\nconsole.log(JSON.stringify({META[lang]['export']}));\n")
        path = f.name
    return json.loads(subprocess.check_output(["node", path]))


def build_xlsx(lang, cats, links):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    m = META[lang]
    wb = Workbook()

    info = wb.active
    info.title = "Info" if lang == "en" else "はじめに"
    info.column_dimensions["A"].width = 110
    rows = [
        (m["title"], Font(name="Arial", bold=True, size=16)),
        (m["company"], Font(name="Arial", size=11)),
        (m["contact"], Font(name="Arial", size=10)),
        ("", None),
        (f"{m['generated']}: {date.today().isoformat()}", Font(name="Arial", size=10)),
        (m["disclaimer"], Font(name="Arial", size=10)),
        (m["linknote"], Font(name="Arial", size=10)),
    ]
    for i, (text, font) in enumerate(rows, start=1):
        cell = info.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    header_fill = PatternFill("solid", start_color=NAVY)
    header_font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    body_font = Font(name="Arial", size=9)
    link_font = Font(name="Arial", size=9, color=LINK_BLUE, underline="single")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    for cat in cats:
        ws = wb.create_sheet(cat["title"][:31])
        ncols = len(cat["columns"])
        title = ws.cell(row=1, column=1, value=f"{cat['title']} — {len(cat['rows'])} {m['products']}")
        title.font = Font(name="Arial", bold=True, size=12)
        for j, col in enumerate(cat["columns"], start=1):
            c = ws.cell(row=2, column=j, value=col.replace("\n", " "))
            c.font = header_font
            c.fill = header_fill
            c.alignment = wrap_top
        for i, row in enumerate(cat["rows"], start=3):
            for j, val in enumerate(row, start=1):
                c = ws.cell(row=i, column=j, value=val)
                c.font = body_font
                c.alignment = wrap_top
                if j == 1:
                    _, url = oks_link(links, val)
                    if url:
                        c.hyperlink = url
                        c.font = link_font
        note_row = 3 + len(cat["rows"]) + 1
        for k, note in enumerate(cat["notes"]):
            c = ws.cell(row=note_row + k, column=1, value=note)
            c.font = Font(name="Arial", size=8, italic=True)
            c.alignment = wrap_top
            ws.merge_cells(start_row=note_row + k, start_column=1, end_row=note_row + k, end_column=ncols)
        widths = {1: 26, 2: 60}
        for j in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(j)].width = widths.get(j, 14)
        ws.freeze_panes = "A3"

    out = OUT / f"factory-express-catalogue-{lang}.xlsx"
    wb.save(out)
    return out


def build_pdf(lang, cats, links):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    if lang == "ja":
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont

        pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
        font = "HeiseiKakuGo-W5"
    else:
        font = "Helvetica"

    m = META[lang]
    navy = colors.HexColor(f"#{NAVY}")
    cell = ParagraphStyle("cell", fontName=font, fontSize=6.5, leading=8.5)
    head = ParagraphStyle("head", parent=cell, textColor=colors.white)
    note = ParagraphStyle("note", parent=cell, fontSize=6, leading=8, textColor=colors.HexColor("#444444"))
    h1 = ParagraphStyle("h1", fontName=font, fontSize=18, leading=24)
    h2 = ParagraphStyle("h2", fontName=font, fontSize=12, leading=16, textColor=navy)
    meta_style = ParagraphStyle("meta", fontName=font, fontSize=8, leading=11)

    out = OUT / f"factory-express-catalogue-{lang}.pdf"
    doc = SimpleDocTemplate(
        str(out),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=m["title"],
        author=m["company"],
    )
    avail = doc.width

    def esc(s):
        return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br/>")

    story = [
        Paragraph(esc(m["title"]), h1),
        Spacer(1, 4),
        Paragraph(esc(f"{m['company']} — {m['contact']}"), meta_style),
        Paragraph(esc(f"{m['generated']}: {date.today().isoformat()}"), meta_style),
        Paragraph(esc(m["disclaimer"]), meta_style),
        Paragraph(esc(m["linknote"]), meta_style),
        Spacer(1, 10),
    ]

    for ci, cat in enumerate(cats):
        ncols = len(cat["columns"])
        if ncols == 10:  # greases / pastes: name, features, then 8 narrow spec cols
            widths = [0.14, 0.24] + [0.62 / 8] * 8
        elif ncols == 8:
            widths = [0.16, 0.32] + [0.52 / 6] * 6
        else:  # related products: 7 cols
            widths = [0.08, 0.16, 0.22, 0.12, 0.10, 0.10, 0.22]
        col_widths = [w * avail for w in widths]

        def name_para(v):
            prefix, url = oks_link(links, v)
            if not url:
                return Paragraph(esc(v), cell)
            markup = f'<link href="{url}" color="#{LINK_BLUE}"><u>{esc(prefix)}</u></link>{esc(v[len(prefix):])}'
            return Paragraph(markup, cell)

        data = [[Paragraph(esc(c), head) for c in cat["columns"]]]
        for row in cat["rows"]:
            data.append([name_para(row[0])] + [Paragraph(esc(v), cell) for v in row[1:]])

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), navy),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5F9")]),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C9D2DD")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                ]
            )
        )
        story.append(Paragraph(esc(f"{ci + 1:02d}  {cat['title']} — {len(cat['rows'])} {m['products']}"), h2))
        story.append(Spacer(1, 4))
        story.append(table)
        story.append(Spacer(1, 4))
        for n in cat["notes"]:
            story.append(Paragraph(esc(n), note))
        if ci < len(cats) - 1:
            story.append(PageBreak())

    doc.build(story)
    return out


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    links = load_oks_links()
    for lang in ("ja", "en"):
        cats = load_categories(lang)
        x = build_xlsx(lang, cats, links)
        p = build_pdf(lang, cats, links)
        for f in (x, p):
            print(f"{f.relative_to(ROOT)}  {f.stat().st_size / 1024:.0f}KB")


if __name__ == "__main__":
    main()
